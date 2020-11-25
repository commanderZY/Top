#!/usr/bin/env python
# -*- coding: utf-8 -*-
# @Time    : 2020/11/6 0006 9:24
# @Author  : ZY

import time
import sys
import openpyxl
import pandas as pd
from openpyxl import load_workbook
import matplotlib.pyplot as plt
from openpyxl.drawing.image import Image


def add_sheet(data, excel_writer, sheet_name):
    """
    不改变原有Excel的数据，新增sheet。
    注：
        使用openpyxl操作Excel时Excel必需存在，因此要新建空sheet
        无论如何sheet页都会被新建，只是当sheet_name已经存在时会新建一个以1结尾的sheet，如：test已经存在时，新建sheet为test1，以此类推
    :param data: DataFrame数据
    :param excel_writer: 文件路径
    :param sheet_name: 新增的sheet名称
    :return:
    """
    book = load_workbook(excel_writer.path)
    excel_writer.book = book
    data.to_excel(excel_writer=excel_writer, sheet_name=sheet_name, index=None, header=True)

    excel_writer.close()


# pandas读取top文件
if len(sys.argv) == 2:
    original_data = pd.read_csv(sys.argv[1], sep='\\s+', encoding='utf-8')
else:
    print("参数个数错误，请指定解析文件")
    sys.exit(-1)

# 以当前时间创建文件名
path = './' + 'top数据分析结果_' + time.strftime("%Y%m%d%H%M%S") + '.xlsx'
print(path)

# 创建空excel，添加sheet的接口必须文件存在
excel_write = pd.ExcelWriter(path, engine='openpyxl')
pd.DataFrame().to_excel(path, sheet_name='Sheet1')

# 可能存储的top文件没有线程调度策略这一项
if original_data.shape[1] == 10:
    original_data.columns = ['进程id(PID)', '进程优先级(PR)', 'CPU占用率(CPU%)', '进程状态(S)', '线程数(#THR)',
                             '虚拟内存(VSS)', '实际使用的物理内存(RSS)', '线程调度策略(PCY)', '进程所有者的ID(UID)', '进程名字(Name)']
elif original_data.shape[1] == 9:
    original_data.columns = ['进程id(PID)', '进程优先级(PR)', 'CPU占用率(CPU%)', '进程状态(S)', '线程数(#THR)',
                             '虚拟内存(VSS)', '实际使用的物理内存(RSS)', '进程所有者的ID(UID)', '进程名字(Name)']
print(original_data)

# 删除有空值的一行，即取出最后一列进程名字不为空的数据
topdata = original_data[original_data['进程名字(Name)'].notna()].copy()
print(topdata)

# 去除CPU占用中的%和内存中的K
topdata['CPU占用率(CPU%)'].replace(regex=True, inplace=True, to_replace='%', value='')
topdata['实际使用的物理内存(RSS)'].replace(regex=True, inplace=True, to_replace='K', value='')
# 转换类型为int/float
topdata['CPU占用率(CPU%)'] = topdata['CPU占用率(CPU%)'].astype('float64')
topdata['实际使用的物理内存(RSS)'] = topdata['实际使用的物理内存(RSS)'].astype('float64')

# 计算统计结果
min_mem = topdata['实际使用的物理内存(RSS)'].min()
max_mem = topdata['实际使用的物理内存(RSS)'].max()
avg_mem = round(topdata['实际使用的物理内存(RSS)'].mean(), 2)
statistical_res = {'统计结果': ['总行数:' + str(len(topdata)),
                            'ky_stb PID:' + str(topdata.iat[0, 0]),
                            '运行时长:' + str(round(len(topdata) * 3 / 60 / 60, 2)) + 'h',
                            'ky_stb重启次数:' + str(topdata['进程id(PID)'].value_counts().count() - 1),
                            '最小CPU:' + str(topdata['CPU占用率(CPU%)'].min()) + '%',
                            '最大CPU:' + str(topdata['CPU占用率(CPU%)'].max()) + '%',
                            '平均CPU:' + str(round(topdata['CPU占用率(CPU%)'].mean(), 2)) + '%',
                            '最小内存:' + str(min_mem) + 'KB' + '(' + str(round(min_mem / 1024, 2)) + 'M)',
                            '最大内存:' + str(max_mem) + 'KB' + '(' + str(round(max_mem / 1024, 2)) + 'M)',
                            '平均内存:' + str(avg_mem) + 'KB' + '(' + str(round(avg_mem / 1024, 2)) + 'M)']}

# 将内存数据转换为折线图并存储为图片
mem_show = pd.DataFrame(topdata['实际使用的物理内存(RSS)'], index=topdata.index)
fig = plt.figure()
# 设置字体，防止默认字体不存在告警以及显示方块
plt.rcParams['font.sans-serif'] = ['SimHei']
ax = fig.add_subplot(1, 1, 1)
mem_show.plot(ax=ax)
fig.savefig('mem.png')

# 将统计结果放入新sheet中
add_sheet(pd.DataFrame(statistical_res), excel_write, sheet_name='统计结果')
# 添加清洗后的未分析数据到excel的新sheet中
add_sheet(pd.DataFrame(topdata), excel_write, sheet_name='原始top数据清洗结果')
# 添加最原始的数据
add_sheet(pd.DataFrame(original_data), excel_write, sheet_name='原始top数据')
# 保存之前的修改
excel_write.save()

# 删除创建时的空sheet
excel_file = path
wb = openpyxl.load_workbook(excel_file)
ws = wb['Sheet1']
wb.remove(ws)
# 设置统计结果列的列宽
ws = wb[wb.sheetnames[0]]
ws.column_dimensions['A'].width = 50
# 插入内存处理后的图片到D3位置
img = Image('mem.png')
ws.add_image(img, 'D3')

wb.save(excel_file)
